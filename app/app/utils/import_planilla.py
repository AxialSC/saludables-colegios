"""
app/utils/import_planilla.py — Lector de la planilla del mayorista.

Soporta .xlsx (directo, sin convertir) y .csv (separador ; o ,).
Detecta SOLO en que fila esta el encabezado y en que columnas estan
Rubro / Codigo / Nombre / Pcio, asi no se rompe si el archivo trae
una columna vacia al principio o cambia un poco de lugar.

La funcion principal devuelve una lista de dicts:
    {'rubro': str, 'codigo': str, 'nombre': str, 'costo_neto': float}
"""
import csv
import unicodedata


def _normalizar(texto):
    """Pasa a minusculas y saca acentos/saltos de linea, para comparar encabezados."""
    if texto is None:
        return ''
    t = str(texto).strip().lower().replace('\n', ' ').replace('\r', ' ')
    t = ''.join(c for c in unicodedata.normalize('NFD', t)
                if unicodedata.category(c) != 'Mn')
    return ' '.join(t.split())


def parse_precio(valor):
    """
    Convierte el precio a float.
    Acepta numero directo o texto tipo '1.053,233' (punto miles, coma decimal).
    Devuelve None si no se puede.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    t = str(valor).strip()
    if not t:
        return None
    # Saca simbolos de moneda y espacios
    t = t.replace('$', '').replace(' ', '')
    # Formato argentino: punto = miles, coma = decimal
    t = t.replace('.', '').replace(',', '.')
    try:
        return float(t)
    except ValueError:
        return None


def _detectar_columnas(fila):
    """
    Dada una fila (lista de celdas), intenta ubicar las columnas.
    Devuelve un dict {'rubro': idx, 'codigo': idx, 'nombre': idx, 'pcio': idx}
    o None si no parece un encabezado.
    """
    mapa = {}
    for idx, celda in enumerate(fila):
        n = _normalizar(celda)
        if not n:
            continue
        if 'rubro' in n and 'rubro' not in mapa:
            mapa['rubro'] = idx
        elif 'codigo' in n and 'codigo' not in mapa:
            mapa['codigo'] = idx
        elif 'nombre' in n and 'nombre' not in mapa:
            mapa['nombre'] = idx
        elif ('pcio' in n or 'precio' in n) and 'pcio' not in mapa:
            mapa['pcio'] = idx
    # Necesitamos al menos codigo, nombre y precio para que sirva
    if all(k in mapa for k in ('codigo', 'nombre', 'pcio')):
        return mapa
    return None


def _filas_xlsx(ruta):
    from openpyxl import load_workbook
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    for fila in ws.iter_rows(values_only=True):
        yield list(fila)
    wb.close()


def _filas_csv(ruta):
    # Probamos separador ; (el habitual del mayorista) y si no, coma
    with open(ruta, 'r', encoding='utf-8-sig', newline='') as f:
        muestra = f.read(2048)
        f.seek(0)
        sep = ';' if muestra.count(';') >= muestra.count(',') else ','
        for fila in csv.reader(f, delimiter=sep):
            yield fila


def leer_planilla(ruta):
    """
    Lee el archivo y devuelve (productos, resumen_lectura).
      productos -> lista de dicts {rubro, codigo, nombre, costo_neto}
      resumen   -> dict con info util para mostrar / diagnosticar
    Lanza ValueError si no encuentra el encabezado o no hay datos.
    """
    ruta_baja = ruta.lower()
    if ruta_baja.endswith(('.xlsx', '.xlsm')):
        generador = _filas_xlsx(ruta)
    elif ruta_baja.endswith('.csv'):
        generador = _filas_csv(ruta)
    else:
        raise ValueError('Formato no soportado. Subí un archivo .xlsx o .csv')

    columnas = None
    productos = []
    descartadas = 0
    total_filas = 0

    for fila in generador:
        total_filas += 1

        # Buscamos el encabezado en las primeras filas
        if columnas is None:
            columnas = _detectar_columnas(fila)
            continue  # la fila de encabezado no es dato

        # Ya tenemos columnas: leemos datos
        def celda(clave):
            idx = columnas.get(clave)
            if idx is None or idx >= len(fila):
                return None
            return fila[idx]

        rubro = celda('rubro')
        codigo = celda('codigo')
        nombre = celda('nombre')
        pcio = celda('pcio')

        # Fila totalmente vacia -> la salteamos sin contar como error
        if not any([rubro, codigo, nombre, pcio]):
            continue

        costo = parse_precio(pcio)
        codigo = str(codigo).strip() if codigo is not None else ''
        nombre = str(nombre).strip() if nombre is not None else ''
        rubro = str(rubro).strip().upper() if rubro else 'SIN RUBRO'

        # Validacion minima: codigo, nombre y costo > 0
        if not codigo or not nombre or not costo or costo <= 0:
            descartadas += 1
            continue

        productos.append({
            'rubro': rubro,
            'codigo': codigo,
            'nombre': nombre,
            'costo_neto': round(costo, 3),
        })

    if columnas is None:
        raise ValueError(
            'No encontré el encabezado (Rubro / Código / Nombre / Pcio). '
            'Verificá que el archivo sea la planilla del mayorista.')

    if not productos:
        raise ValueError('No se detectó ningún producto válido en el archivo.')

    resumen = {
        'total_filas': total_filas,
        'productos_leidos': len(productos),
        'descartadas': descartadas,
        'columnas_detectadas': columnas,
    }
    return productos, resumen
